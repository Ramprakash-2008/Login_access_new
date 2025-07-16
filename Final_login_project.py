from flask import Flask, request, redirect, url_for, render_template_string, session, send_file
import sqlite3, os, smtplib
from datetime import datetime
from email.mime.text import MIMEText
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
from datetime import datetime, time
import pytz
from dotenv import load_dotenv
from flask import flash  
import logging
import requests
from geopy.geocoders import Nominatim
load_dotenv()
app = Flask(__name__)
app.secret_key = "super_secret_key"
 
# === CONFIG ===
ADMIN_PASSWORD = os.getenv("ADMIN_PASS")
ADMIN_EMAIL = os.getenv("EMAIL_USER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASS")
          # <-- Replace this
DATABASE = "users.db"
LOG_FILE = "login_log.csv"
DEADLINE_TIME = "09:00"

# === HTML Templates ===
LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang='en'>
<head>
    <meta charset='UTF-8'>
    <meta name='viewport' content='width=device-width, initial-scale=1.0'>
    <title>Admin Login</title>
    <script>
     if (location.protocol !== 'https:') {
       location.href = 'https:' + window.location.href.substring(window.location.protocol.length);   
     }
    </script>

    <style>
        body {
            background-color: #fafafa;
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
            font-family: 'Segoe UI', sans-serif;
        }
        .login-box {
            background: white;
            border: 1px solid #dbdbdb;
            padding: 40px;
            width: 350px;
            text-align: center;
        }
        h2 {
            font-size: 30px;
            margin-bottom: 20px;
        }
        input[type="password"] {
            width: 100%;
            padding: 10px;
            margin-bottom: 10px;
            border: 1px solid #ccc;
            border-radius: 4px;
        }
        button {
            width: 100%;
            padding: 10px;
            background-color: #3897f0;
            color: white;
            border: none;
            border-radius: 4px;
            font-weight: bold;
            cursor: pointer;
        }
        a {
            display: block;
            margin-top: 15px;
            font-size: 14px;
            color: #00376b;
            text-decoration: none;
        }
    </style>
</head>
<body>
    <form method="POST">
        <div class="login-box">
            <h2>Admin</h2>
            <input type="password" name="password" placeholder="Password" required>
            <button type="submit">Log In</button>
            <a href="/">← Back to Login</a>
        </div>
    </form>
</body>
</html>

"""

ADMIN_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Admin Dashboard</title>
    <style>
        body { font-family: Arial; background: #f4f4f4; margin: 0; padding: 20px; }
        .container { max-width: 1000px; margin: auto; background: #fff; padding: 20px; border-radius: 8px; }
        .topbar button { margin-right: 10px; }
        .hidden { display: none; }
        .status { color: green; font-weight: bold; }
        .late { color: red; }
        .ontime { color: green; }
        table { width: 100%; border-collapse: collapse; }
        table, th, td { border: 1px solid #ddd; }
        th, td { padding: 10px; text-align: center; }
        button { padding: 10px; margin-top: 10px; border-radius: 5px; border: none; background: #3897f0; color: #fff; cursor: pointer; }
        button:hover { background: #287bd1; }
        h2, h3 { text-align: center; }
    </style>
    <script>
        function toggle(id) {
            const el = document.getElementById(id);
            el.style.display = (el.style.display === "none") ? "block" : "none";
        }
    </script>
</head>
<body>
<div class="container">
<h2>Admin Dashboard - Login Logs</h2>
<div class="topbar">
    <button onclick="toggle('addUserForm')">➕ Add User</button>
    <button onclick="toggle('updateUsersForm')">🛠 Update Users</button>
    <form method="GET" action="/users" style="display:inline;">
        <button>👤 Show All Registered Users</button>
    </form>
    <form method="GET" action="/download-log" style="display:inline;">
        <button>📥 Download Excel Log</button>
    </form>
        <form method="GET" action="/not-logged-in" style="display:inline;">
        <button>🙋‍♂️ View Not Logged In Users</button>
    </form>
    <form method="GET" action="/clear-log" style="display:inline;">
        <button>🧹 Clear All Logs</button>
    </form>

    <form method="GET" action="/logout" style="display:inline;">
        <button>🔒 Logout</button>
    </form>
</div>

{% if message %}<p class="status">✅ {{ message }}</p>{% endif %}

<div id="addUserForm" class="hidden">
    <form method="POST">
        <label>Add Single Username:</label>
        <input type="text" name="new_username" required>
        <button type="submit">Add</button>
    </form>
    <br><hr>
    <form method="POST" enctype="multipart/form-data">
        <label>Upload Usernames File (.txt/.csv):</label>
        <input type="file" name="file" accept=".txt,.csv" required>
        <button type="submit">Upload</button>
    </form>
</div>

<div id="updateUsersForm" class="hidden">
    <form method="POST">
        <table>
            <tr><th>Username</th><th>Edit</th><th>Delete</th></tr>
            {% for user in users %}
            <tr>
                <td><input type="text" name="usernames" value="{{ user }}"></td>
                <td><input type="checkbox" name="edit_{{ loop.index0 }}"></td>
                <td><input type="checkbox" name="delete_{{ loop.index0 }}"></td>
                <input type="hidden" name="original_{{ loop.index0 }}" value="{{ user }}">
            </tr>
            {% endfor %}
        </table>
        <input type="hidden" name="count" value="{{ users|length }}">
        <button type="submit" name="save_changes">💾 Save Changes</button>
    </form>
</div>

<h3>Login Logs</h3>
<table>
<tr>
    <th>Username</th><th>Time</th><th>Status</th><th>Date</th><th>Location</th>
</tr>
{% for row in logs %}
<tr>
    <td>{{ row.username }}</td>
    <td>{{ row.time }}</td>
    <td class="{{ 'late' if row.status == 'Late' else 'ontime' }}">{{ row.status }}</td>
    <td>{{ row.date }}</td>
    <td>{{ row.location }}</td>
</tr>
{% endfor %}
</table>

</div>
</body>
</html>

"""

USERS_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>All Registered Users</title>
    <style>
        body { font-family: Arial; background: #fafafa; padding: 20px; text-align: center; }
        ul { list-style-type: none; padding: 0; }
        li { padding: 5px; font-size: 18px; }
        a { text-decoration: none; color: #3897f0; font-weight: bold; }
    </style>
</head>
<body>
    <h2>All Registered Users</h2>
    <ul>
    {% for user in users %}
        <li>{{ user }}</li>
    {% endfor %}
    </ul>
    <a href="/admin">⬅ Back to Admin</a>
</body>
</html>

"""

# === Helper Functions ===
def init_db():
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS users (username TEXT UNIQUE)")
    conn.commit()
    conn.close()
init_db()
def get_users():
    conn = sqlite3.connect(DATABASE)
    cur = conn.cursor()
    cur.execute("SELECT username FROM users")
    users = [row[0] for row in cur.fetchall()]
    conn.close()
    return users
def get_address(lat, lng):
    try:
        if not lat or not lng:
            return "Unknown"
        url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lng}"
        headers = {"User-Agent": "FlaskLoginTracker/1.0"}
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            return response.json().get("display_name", f"{lat},{lng}")
        return f"{lat},{lng}"
    except Exception as e:
        print("⚠️ Reverse geocoding failed:", e)
        return f"{lat},{lng}"

def save_log(username, time, status, date, location):
    if not os.path.exists(LOG_FILE):
        with open(LOG_FILE, 'w') as f:
            f.write("Username,Time,Status,Date,Location\n")
    with open(LOG_FILE, 'a') as f:
        f.write(f"{username},{time},{status},{date},{location}\n")



logging.basicConfig(level=logging.WARNING, format='%(levelname)s: %(message)s')

import logging

def get_logs():
    if not os.path.exists(LOG_FILE):
        return []

    logs = []
    with open(LOG_FILE, 'r') as f:
        next(f, None)  # Skip header
        for line in f:
            parts = line.strip().split(",")
            if len(parts) >= 5:
                username = parts[0]
                time = parts[1]
                status = parts[2]
                date = parts[3]
                location = ",".join(parts[4:])  # Allow commas in address
                logs.append([username, time, status, date, location])
            else:
                logging.warning("⚠️ Skipping malformed line: %s", line.strip())
    return logs


def send_late_email(username, login_time):
    msg = MIMEText(f"User '{username}' logged in late at {login_time}.")
    msg['Subject'] = f"LATE LOGIN: {username}"
    msg['From'] = ADMIN_EMAIL
    msg['To'] = ADMIN_EMAIL
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(ADMIN_EMAIL, EMAIL_PASSWORD)
            server.sendmail(ADMIN_EMAIL, ADMIN_EMAIL, msg.as_string())
    except Exception as e:
        print("❌ Email failed:", e)



@app.route("/download-log")
def download_excel():
    if not session.get("logged_in"):
        return redirect("/admin/login")

    if not os.path.exists(LOG_FILE):
        return "Log file not found."

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Login Log"
        ws.append(["Username", "Time", "Status", "Date", "Location"])

        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        with open(LOG_FILE, "r") as f:
            next(f)  # skip header
            for line in f:
                parts = line.strip().split(",")
                if len(parts) >= 5:
                    username = parts[0]
                    log_time = parts[1]
                    status = parts[2]
                    log_date = parts[3]
                    location = ",".join(parts[4:])  # to handle address commas
                    row = [username, log_time, status, log_date, location]
                    ws.append(row)
                    fill = green if status == "On-time" else red
                    for cell in ws[ws.max_row]:
                        cell.fill = fill
                else:
                    print("⚠️ Skipping malformed line:", line.strip())

        output_folder = os.path.join(os.path.dirname(__file__), "downloads")
        os.makedirs(output_folder, exist_ok=True)
        file_path = os.path.join(output_folder, "login_log.xlsx")

        wb.save(file_path)
        return send_file(file_path, as_attachment=True)

    except Exception as e:
        print("❌ Error generating Excel:", e)
        return f"❌ Error generating Excel: {e}", 500



@app.route("/not-logged-in")
def not_logged_in_users():
    if not session.get("logged_in"):
        return redirect("/admin/login")

    today_users = set()
    if os.path.exists(LOG_FILE):
        with open(LOG_FILE) as f:
            next(f)
            for line in f:
                parts = line.strip().split(',')
                if len(parts) == 3:
                    today_users.add(parts[0])

    registered_users = set(get_users())
    not_logged_in = sorted(list(registered_users - today_users))

    return render_template_string("""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Users Not Logged In</title>
        <style>
            body {
                background-color: #fafafa;
                font-family: 'Segoe UI', sans-serif;
                display: flex;
                justify-content: center;
                align-items: center;
                flex-direction: column;
                padding: 40px;
            }
            .box {
                background: white;
                border: 1px solid #dbdbdb;
                padding: 40px;
                width: 400px;
                text-align: center;
            }
            h2 {
                margin-bottom: 20px;
            }
            ul {
                list-style: none;
                padding: 0;
                text-align: left;
            }
            li {
                padding: 5px 0;
                border-bottom: 1px solid #eee;
                font-size: 16px;
            }
            a {
                display: inline-block;
                margin-top: 20px;
                font-size: 14px;
                color: #3897f0;
                text-decoration: none;
            }
        </style>
    </head>
    <body>
        <div class="box">
            <h2>🙋‍♂️ Users Not Logged In Today</h2>
            {% if users %}
            <ul>
                {% for user in users %}
                    <li>{{ user }}</li>
                {% endfor %}
            </ul>
            {% else %}
            <p>✅ All users have logged in!</p>
            {% endif %}
            <a href="/admin">⬅ Back to Admin</a>
        </div>
    </body>
    </html>
    """, users=not_logged_in)



@app.route("/clear-log")
def clear_log():
    if not session.get("logged_in"):
        return redirect("/admin/login")

    with open(LOG_FILE, "w") as f:
        f.write("Username,Time,Status\n")  # Keep header only

    return redirect("/admin?message=✅ Logs Cleared Successfully")

@app.route("/")
def home():
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <title>User Login</title>
        <style>
            body {
                background-color: #fafafa;
                font-family: 'Segoe UI', sans-serif;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
            }
            .login-box {
                background: white;
                border: 1px solid #dbdbdb;
                padding: 40px;
                text-align: center;
                width: 350px;
            }
            input, button {
                width: 100%;
                padding: 10px;
                margin: 10px 0;
                border-radius: 4px;
                border: 1px solid #ccc;
            }
            button {
                background-color: #3897f0;
                color: white;
                font-weight: bold;
                cursor: pointer;
                border: none;
            }
            a {
                display: block;
                margin-top: 10px;
                color: #3897f0;
                text-decoration: none;
            }
        </style>
        <script>
            function sendLogin(event) {
                event.preventDefault();
                const username = document.getElementById('username').value;

                if (!username) {
                    alert("Please enter a username.");
                    return;
                }

                if (navigator.geolocation) {
                    navigator.geolocation.getCurrentPosition(function(position) {
                        const lat = position.coords.latitude;
                        const lng = position.coords.longitude;

                        fetch("/login", {
                            method: "POST",
                            headers: {
                                "Content-Type": "application/json"
                            },
                            body: JSON.stringify({
                                username: username,
                                lat: lat,
                                lng: lng
                            })
                        })
                        .then(res => res.text())
                        .then(html => {
                            document.open();
                            document.write(html);
                            document.close();
                        })
                        .catch(err => alert("Login failed: " + err));
                    }, function(error) {
                        alert("Geolocation permission denied or unavailable.");
                    });
                } else {
                    alert("Geolocation not supported.");
                }
            }
        </script>
    </head>
    <body>
        <div class="login-box">
            <h2>User Login</h2>
            <form onsubmit="sendLogin(event)">
                <input type="text" id="username" placeholder="Enter username" required />
                <button type="submit">Log In</button>
            </form>
            <a href="/register">📝 Register</a>
            <a href="/admin/login">🔐 Admin Login</a>
        </div>
    </body>
    </html>
    ''')

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        if not username:
            return "❌ Username required", 400

        conn = sqlite3.connect(DATABASE)
        cur = conn.cursor()
        cur.execute("SELECT username FROM users WHERE username=?", (username,))
        exists = cur.fetchone()

        if exists:
            return render_template_string(f'''
            <!DOCTYPE html>
            <html>
            <head>
                <title>User Exists</title>
                <style>
                    body {{
                        background-color: #fafafa;
                        display: flex;
                        justify-content: center;
                        align-items: center;
                        height: 100vh;
                        font-family: 'Segoe UI', sans-serif;
                    }}
                    .login-box {{
                        background: white;
                        border: 1px solid #dbdbdb;
                        padding: 40px;
                        width: 350px;
                        text-align: center;
                    }}
                    h2 {{ color: red; }}
                    a {{
                        display: block;
                        margin-top: 15px;
                        font-size: 14px;
                        color: #00376b;
                        text-decoration: none;
                    }}
                </style>
            </head>
            <body>
                <div class="login-box">
                    <h2>❌ User Already Exists</h2>
                    <a href="/">← Back to Login</a>
                </div>
            </body>
            </html>
            ''')
        else:
            cur.execute("INSERT INTO users (username) VALUES (?)", (username,))
            conn.commit()
            conn.close()
            return render_template_string(f'''
            <!DOCTYPE html>
            <html>
            <head>
                <title>Registered</title>
                <style>
                    body {{
                        background-color: #fafafa;
                        display: flex;
                        justify-content: center;
                        align-items: center;
                        height: 100vh;
                        font-family: 'Segoe UI', sans-serif;
                    }}
                    .login-box {{
                        background: white;
                        border: 1px solid #dbdbdb;
                        padding: 40px;
                        width: 350px;
                        text-align: center;
                    }}
                    h2 {{ color: green; }}
                    a {{
                        display: block;
                        margin-top: 15px;
                        font-size: 14px;
                        color: #00376b;
                        text-decoration: none;
                    }}
                </style>
            </head>
            <body>
                <div class="login-box">
                    <h2>✅ Successfully Registered</h2>
                    <a href="/">← Back to Login</a>
                </div>
            </body>
            </html>
            ''')
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Register</title>
        <style>
            body {
                background-color: #fafafa;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
                font-family: 'Segoe UI', sans-serif;
            }
            .login-box {
                background: white;
                border: 1px solid #dbdbdb;
                padding: 40px;
                width: 350px;
                text-align: center;
            }
            h2 {
                font-size: 28px;
                margin-bottom: 20px;
            }
            input[type="text"] {
                width: 100%;
                padding: 10px;
                margin-bottom: 10px;
                border: 1px solid #ccc;
                border-radius: 4px;
            }
            button {
                width: 100%;
                padding: 10px;
                background-color: #3897f0;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                cursor: pointer;
            }
            a {
                display: block;
                margin-top: 15px;
                font-size: 14px;
                color: #00376b;
                text-decoration: none;
            }
        </style>
    </head>
    <body>
        <form method="POST">
            <div class="login-box">
                <h2>Register</h2>
                <input type="text" name="username" placeholder="Choose username" required>
                <button type="submit">Register</button>
                <a href="/">← Back to Login</a>
            </div>
        </form>
    </body>
    </html>
    ''')


@app.route("/login", methods=["POST"])
def login_user():
    try:
        # Try both form and JSON (fallback)
        if request.is_json:
            data = request.get_json()
            username = data.get("username", "").strip()
            lat = data.get("lat")
            lng = data.get("lng")
        else:
            username = request.form.get("username", "").strip()
            lat = request.form.get("lat")
            lng = request.form.get("lng")

        if not username:
            return "❌ Username required", 400

        # Validate user
        conn = sqlite3.connect(DATABASE)
        c = conn.cursor()
        c.execute("SELECT username FROM users WHERE username=?", (username,))
        result = c.fetchone()
        conn.close()

        if not result:
            return render_template_string("""
            <h2 style="color:red;">❌ You are not a registered user</h2>
            <a href="/">← Back to Login</a>
            """)

        # Get time, date, location
        india = pytz.timezone("Asia/Kolkata")
        now = datetime.now(india)
        current_time = now.time()
        deadline = time(hour=9, minute=0)
        status = "Late" if current_time > deadline else "On-time"
        log_time = now.strftime("%H:%M")
        log_date = now.strftime("%Y-%m-%d")

        # Convert lat/lng to human-readable location
        location = "Unknown"
        if lat and lng:
            try:
                url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lng}"
                headers = {"User-Agent": "FlaskLoginTracker/1.0"}  # required by Nominatim
                response = requests.get(url, headers=headers, timeout=10)
                if response.status_code == 200:
                    data = response.json()
                    location = data.get("display_name", f"{lat},{lng}")
                else:
                    print(f"⚠️ Nominatim error {response.status_code}: {response.text}")
                    location = f"{lat},{lng}"
            except Exception as geo_err:
                print("⚠️ Reverse geocoding failed:", geo_err)
                location = f"{lat},{lng}"


        # Save to log
        save_log(username, log_time, status, log_date, location)

        # Send email if late
        if status == "Late":
            send_late_email(username, log_time)

        return render_template_string(f'''
        <html><head><title>Login Success</title>
        <style>
        body {{
            background-color: #fafafa;
            display: flex; justify-content: center; align-items: center;
            height: 100vh; font-family: 'Segoe UI', sans-serif;
        }}
        .login-box {{
            background: white; padding: 40px; text-align: center;
            border: 1px solid #dbdbdb; width: 350px;
        }}
        </style>
        </head><body>
        <div class="login-box">
            <h2>✅ Login Successful</h2>
            <p><b>{username}</b> logged in at <b>{log_time}</b></p>
            <p>Status: <span style="color:{'red' if status == 'Late' else 'green'};">{status}</span></p>
            <p><small>{log_date}</small></p>
            <a href="/">← Back to Login</a>
        </div>
        </body></html>
        ''')

    except Exception as e:
        print("❌ ERROR in /login:", e)
        return "Internal Server Error", 500



@app.route("/admin", methods=["GET", "POST"])
def admin():
    if not session.get("logged_in"):
        return redirect("/admin/login")

    message = ""

    if request.method == "POST":
        if "new_username" in request.form:
            username = request.form["new_username"].strip()
            if username:
                conn = sqlite3.connect(DATABASE)
                cur = conn.cursor()
                cur.execute("INSERT OR IGNORE INTO users (username) VALUES (?)", (username,))
                conn.commit()
                conn.close()
                message = f"User '{username}' added."

        elif "file" in request.files:
            file = request.files["file"]
            if file:
                lines = file.read().decode("utf-8").splitlines()
                conn = sqlite3.connect(DATABASE)
                cur = conn.cursor()
                for name in lines:
                    cur.execute("INSERT OR IGNORE INTO users (username) VALUES (?)", (name.strip(),))
                conn.commit()
                conn.close()
                message = f"{len(lines)} users uploaded."

        elif "save_changes" in request.form:
            count = int(request.form["count"])
            conn = sqlite3.connect(DATABASE)
            cur = conn.cursor()
            for i in range(count):
                original = request.form[f"original_{i}"].strip()
                new_value = request.form.getlist("usernames")[i].strip()
                delete = request.form.get(f"delete_{i}")

                if delete:
                    cur.execute("DELETE FROM users WHERE username = ?", (original,))
                elif new_value != original:
                    cur.execute("UPDATE users SET username = ? WHERE username = ?", (new_value, original))
            conn.commit()
            conn.close()
            message = "Changes saved successfully."

    # ✅ Pass logs with 5 fields
    logs = []
    for row in get_logs():
        if len(row) == 5:
            logs.append({
                "username": row[0],
                "time": row[1],
                "status": row[2],
                "date": row[3],
                "location": row[4],
            })
        else:
            print("⚠️ Skipping malformed log line:", row)

    return render_template_string(ADMIN_TEMPLATE, logs=logs, users=get_users(), message=message)


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if request.form["password"] == ADMIN_PASSWORD:
            session["logged_in"] = True
            return redirect("/admin")
        else:
            return render_template_string("""
                <h2 style=\"color:red;\">❌ Incorrect Password</h2>
                <a href=\"/admin/login\">🔁 Try Again</a>
            """)
    return render_template_string(LOGIN_TEMPLATE)

@app.route("/users")
def show_users():
    return render_template_string(USERS_TEMPLATE, users=get_users())

@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect("/admin/login")

# === Main ===
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
